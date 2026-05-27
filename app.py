import io
import re
import zipfile
from datetime import datetime
from pathlib import Path
from xml.etree import ElementTree as ET

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="US Foods Report Tool", page_icon="📦", layout="wide")

st.title("📦 US Foods Report Tool")
st.caption("Keeps past dates, today's date, and the next upcoming ship date only.")

TODAY = pd.Timestamp.today().normalize()

# -----------------------------
# Helpers
# -----------------------------
def clean_text(value):
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def normalize_col(col):
    return re.sub(r"[^a-z0-9]+", "", str(col).lower())


def find_column(df, possible_names):
    normalized = {normalize_col(c): c for c in df.columns}
    for name in possible_names:
        key = normalize_col(name)
        if key in normalized:
            return normalized[key]
    for c in df.columns:
        c_norm = normalize_col(c)
        for name in possible_names:
            if normalize_col(name) in c_norm:
                return c
    return None


def read_excel_any(uploaded_file):
    if uploaded_file is None:
        return None
    name = uploaded_file.name.lower()
    if name.endswith(".xls"):
        return pd.read_excel(uploaded_file, dtype=str, engine="xlrd")
    return pd.read_excel(uploaded_file, dtype=str)


def read_all_xml_texts(uploaded_zips):
    """Reads XML files and text/email body files from one or more zip uploads."""
    records = []
    for uploaded_zip in uploaded_zips or []:
        data = uploaded_zip.read()
        with zipfile.ZipFile(io.BytesIO(data)) as z:
            for filename in z.namelist():
                if filename.endswith("/"):
                    continue
                lower = filename.lower()
                if not lower.endswith((".xml", ".txt", ".eml", ".html", ".htm")):
                    continue
                try:
                    raw = z.read(filename)
                    text = raw.decode("utf-8", errors="ignore")
                    records.append({"source_file": filename, "text": text})
                except Exception:
                    pass
    return records


def xml_value(root, path):
    node = root.find(path)
    return clean_text(node.text if node is not None else "")


def spec_value(root, spec_name):
    for spec in root.findall(".//SPECIFICATION"):
        if clean_text(spec.attrib.get("NAME", "")).lower() == spec_name.lower():
            return clean_text(spec.text)
    return ""


def regex_first(patterns, text, flags=re.I | re.S):
    for pattern in patterns:
        match = re.search(pattern, text, flags)
        if match:
            return clean_text(match.group(1))
    return ""


def parse_requested_date(value):
    value = clean_text(value)
    if not value:
        return pd.NaT
    value = re.sub(r"\b(EST|EDT|CST|CDT|PST|PDT)\b", "", value, flags=re.I).strip()
    return pd.to_datetime(value, errors="coerce")


def parse_order_record(text, source_file=""):
    """Pulls order details from XML when possible, with fallback regex for email body text."""
    rec = {"Source File": source_file}

    try:
        root = ET.fromstring(text.strip())
        rec.update({
            "Order ID": xml_value(root, ".//ORDER_NUMBER"),
            "Customer PO Number": xml_value(root, ".//TP_PO_NUMBER"),
            "Supplier Part ID": xml_value(root, ".//ITEM_CODE"),
            "Order Description": xml_value(root, ".//ITEM_DESCRIPTION"),
            "Quantity": xml_value(root, ".//SHIPPING_QUANTITY") or xml_value(root, ".//QUANTITY"),
            "Requested Delivery Date": xml_value(root, ".//REQUESTED_DELIVERY_DATE") or xml_value(root, ".//REQUEST_DATE"),
            "Paper Type": spec_value(root, "Stock") or spec_value(root, "Paper Type") or spec_value(root, "Paper"),
            "Coating": spec_value(root, "Coating") or "NONE",
            "Page Setup": spec_value(root, "Page Setup"),
            "Color": spec_value(root, "Ink") or spec_value(root, "Color"),
            "Page Size": spec_value(root, "Page Size") or spec_value(root, "Finished Size"),
            "Bindery 1": spec_value(root, "Bindery 1") or spec_value(root, "Bindery"),
            "Bindery 2": spec_value(root, "Bindery 2"),
            "Laminate": spec_value(root, "Laminate") or "NONE",
            "Collate": spec_value(root, "Collate") or "NONE",
        })
        return rec
    except Exception:
        pass

    # Fallback for email bodies that contain XML-like or plain text order info
    rec.update({
        "Order ID": regex_first([r"<ORDER_NUMBER>(.*?)</ORDER_NUMBER>", r"Order\s*(?:ID|Number)\s*[:#-]\s*([^\n\r<]+)"], text),
        "Customer PO Number": regex_first([r"<TP_PO_NUMBER>(.*?)</TP_PO_NUMBER>", r"(?:Customer\s*)?PO\s*(?:Number|#)?\s*[:#-]\s*([^\n\r<]+)"], text),
        "Supplier Part ID": regex_first([r"<ITEM_CODE>(.*?)</ITEM_CODE>", r"Supplier\s*Part\s*ID\s*[:#-]\s*([^\n\r<]+)"], text),
        "Order Description": regex_first([r"<ITEM_DESCRIPTION>(.*?)</ITEM_DESCRIPTION>", r"Order\s*Description\s*[:#-]\s*([^\n\r<]+)"], text),
        "Quantity": regex_first([r"<SHIPPING_QUANTITY>(.*?)</SHIPPING_QUANTITY>", r"Quantity\s*[:#-]\s*([^\n\r<]+)"], text),
        "Requested Delivery Date": regex_first([r"<REQUESTED_DELIVERY_DATE>(.*?)</REQUESTED_DELIVERY_DATE>", r"Requested\s*Delivery\s*Date\s*[:#-]\s*([^\n\r<]+)"], text),
        "Paper Type": regex_first([r'SPECIFICATION NAME="(?:Stock|Paper Type|Paper)">(.*?)</SPECIFICATION>', r"Paper\s*Type\s*[:#-]\s*([^\n\r<]+)"], text),
        "Coating": regex_first([r'SPECIFICATION NAME="Coating">(.*?)</SPECIFICATION>', r"Coating\s*[:#-]\s*([^\n\r<]+)"], text) or "NONE",
        "Page Setup": regex_first([r'SPECIFICATION NAME="Page Setup">(.*?)</SPECIFICATION>', r"Page\s*Setup\s*[:#-]\s*([^\n\r<]+)"], text),
        "Color": regex_first([r'SPECIFICATION NAME="(?:Ink|Color)">(.*?)</SPECIFICATION>', r"Color\s*[:#-]\s*([^\n\r<]+)"], text),
        "Page Size": regex_first([r'SPECIFICATION NAME="(?:Page Size|Finished Size)">(.*?)</SPECIFICATION>', r"Page\s*Size\s*[:#-]\s*([^\n\r<]+)"], text),
        "Bindery 1": regex_first([r'SPECIFICATION NAME="Bindery 1">(.*?)</SPECIFICATION>', r"Bindery\s*1\s*[:#-]\s*([^\n\r<]+)"], text),
        "Bindery 2": regex_first([r'SPECIFICATION NAME="Bindery 2">(.*?)</SPECIFICATION>', r"Bindery\s*2\s*[:#-]\s*([^\n\r<]+)"], text),
        "Laminate": regex_first([r'SPECIFICATION NAME="Laminate">(.*?)</SPECIFICATION>', r"Laminate\s*[:#-]\s*([^\n\r<]+)"], text) or "NONE",
        "Collate": regex_first([r'SPECIFICATION NAME="Collate">(.*?)</SPECIFICATION>', r"Collate\s*[:#-]\s*([^\n\r<]+)"], text) or "NONE",
    })
    return rec


def build_xml_lookup(xml_records):
    parsed = []
    for item in xml_records:
        rec = parse_order_record(item["text"], item["source_file"])
        if rec.get("Customer PO Number") or rec.get("Order ID"):
            parsed.append(rec)
    xml_df = pd.DataFrame(parsed)
    if xml_df.empty:
        return xml_df
    xml_df["PO Base"] = xml_df["Customer PO Number"].astype(str).str.replace(r"-\d+$", "", regex=True).str.strip()
    return xml_df


def add_xml_details(print_df, xml_df):
    if xml_df.empty:
        print_df["Ship Date"] = ""
        return print_df

    po_col = find_column(print_df, ["PO#", "PO Number", "Customer PO Number", "Order #", "PO"])
    if po_col is None:
        st.warning("Could not find a PO column in the Print Logic report.")
        print_df["Ship Date"] = ""
        return print_df

    df = print_df.copy()
    df["_po_exact"] = df[po_col].astype(str).str.strip()
    df["_po_base"] = df["_po_exact"].str.replace(r"-\d+$", "", regex=True).str.strip()

    xml_exact = xml_df.drop_duplicates(subset=["Customer PO Number"], keep="first").set_index("Customer PO Number")
    xml_base = xml_df.drop_duplicates(subset=["PO Base"], keep="first").set_index("PO Base")

    detail_cols = [
        "Order ID", "Requested Delivery Date", "Supplier Part ID", "Paper Type", "Coating",
        "Page Setup", "Color", "Page Size", "Bindery 1", "Bindery 2", "Laminate", "Collate"
    ]

    for col in detail_cols:
        exact_map = df["_po_exact"].map(xml_exact[col]) if col in xml_exact.columns else pd.Series("", index=df.index)
        base_map = df["_po_base"].map(xml_base[col]) if col in xml_base.columns else pd.Series("", index=df.index)
        df[col] = exact_map.fillna(base_map).fillna("")

    df["Ship Date"] = pd.to_datetime(df["Requested Delivery Date"], errors="coerce").dt.strftime("%m/%d/%y")
    df = df.drop(columns=["_po_exact", "_po_base"])
    return df


def remove_shipped_rows(df, tracking_df):
    if tracking_df is None or tracking_df.empty:
        return df

    job_col_df = find_column(df, ["Job No", "Job #", "Job", "Job Number"])
    job_col_tr = find_column(tracking_df, ["Job No", "Job #", "Job", "Job Number"])
    tracking_col = find_column(tracking_df, ["Tracking Number", "Tracking", "Tracking #"])

    if not all([job_col_df, job_col_tr, tracking_col]):
        st.warning("Could not find Job/Tracking columns, so shipped rows were not removed.")
        return df

    shipped_jobs = set(
        tracking_df.loc[tracking_df[tracking_col].astype(str).str.strip().ne(""), job_col_tr]
        .astype(str).str.strip()
    )
    return df[~df[job_col_df].astype(str).str.strip().isin(shipped_jobs)].copy()


def cancelled_job_set(cancelled_df):
    if cancelled_df is None or cancelled_df.empty:
        return set()
    job_col = find_column(cancelled_df, ["Job No", "Job #", "Job", "Job Number"])
    status_col = find_column(cancelled_df, ["Status", "Job Status"])
    if not job_col:
        return set()
    if status_col:
        return set(cancelled_df.loc[cancelled_df[status_col].astype(str).str.contains("cancel", case=False, na=False), job_col].astype(str).str.strip())
    return set(cancelled_df[job_col].astype(str).str.strip())


def apply_ship_date_filter(df):
    ship_col = find_column(df, ["Ship Date", "Requested Delivery Date"])
    if not ship_col:
        st.warning("Could not find Ship Date column, so ship date filtering was skipped.")
        return df

    temp = df.copy()
    temp["_ship_date"] = pd.to_datetime(temp[ship_col], errors="coerce").dt.normalize()

    past_or_today = temp["_ship_date"] <= TODAY
    future_dates = temp.loc[temp["_ship_date"] > TODAY, "_ship_date"].dropna()

    if not future_dates.empty:
        next_ship_date = future_dates.min()
        keep_next = temp["_ship_date"] == next_ship_date
        temp = temp[past_or_today | keep_next].copy()
    else:
        temp = temp[past_or_today].copy()

    temp = temp.sort_values(by="_ship_date", na_position="last")
    temp = temp.drop(columns=["_ship_date"])
    return temp


def add_uv_column(df):
    df = df.copy()
    paper_col = find_column(df, ["Paper Type", "Paper", "Stock"])
    if paper_col:
        df["UV"] = df[paper_col].astype(str).str.contains("UV", case=False, na=False).map({True: "UV", False: ""})
    else:
        df["UV"] = ""
    return df


def split_sheets(df):
    laminate_col = find_column(df, ["Laminate"])
    coating_col = find_column(df, ["Coating"])

    if laminate_col is None:
        df["Laminate"] = "NONE"
        laminate_col = "Laminate"
    if coating_col is None:
        df["Coating"] = "NONE"
        coating_col = "Coating"

    laminate_mask = df[laminate_col].astype(str).str.upper().str.strip().eq("LAMINATE")
    coating_mask = ~df[coating_col].astype(str).str.upper().str.strip().isin(["", "NONE", "NAN"])
    trim_mask = ~laminate_mask & ~coating_mask

    return {
        "All Open Jobs": df,
        "Laminate": df[laminate_mask].copy(),
        "UV COATING": df[coating_mask].copy(),
        "Trim To Size": df[trim_mask].copy(),
    }


def format_workbook(output_bytes, cancelled_jobs):
    wb = load_workbook(output_bytes)
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        headers = [cell.value for cell in ws[1]]
        job_idx = None
        for i, h in enumerate(headers, start=1):
            if normalize_col(h) in ["jobno", "job", "jobnumber", "job#"]:
                job_idx = i
                break

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

            if job_idx and row[0].row > 1:
                job_value = clean_text(ws.cell(row=row[0].row, column=job_idx).value)
                if job_value in cancelled_jobs:
                    for cell in row:
                        cell.fill = red_fill

        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in column_cells:
                max_len = max(max_len, len(clean_text(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 45)

    final = io.BytesIO()
    wb.save(final)
    final.seek(0)
    return final


def create_report(print_logic_file, xml_zip_files, tracking_file=None, cancelled_file=None):
    print_df = read_excel_any(print_logic_file)
    if print_df is None or print_df.empty:
        raise ValueError("Print Logic report is empty or could not be read.")

    tracking_df = read_excel_any(tracking_file) if tracking_file else None
    cancelled_df = read_excel_any(cancelled_file) if cancelled_file else None

    xml_records = read_all_xml_texts(xml_zip_files)
    xml_df = build_xml_lookup(xml_records)

    df = add_xml_details(print_df, xml_df)
    df = remove_shipped_rows(df, tracking_df)
    df = add_uv_column(df)
    df = apply_ship_date_filter(df)

    cancelled_jobs = cancelled_job_set(cancelled_df)
    sheets = split_sheets(df)

    raw_output = io.BytesIO()
    with pd.ExcelWriter(raw_output, engine="openpyxl") as writer:
        for sheet_name, sheet_df in sheets.items():
            safe_name = sheet_name[:31]
            sheet_df.to_excel(writer, index=False, sheet_name=safe_name)
    raw_output.seek(0)

    return format_workbook(raw_output, cancelled_jobs), df, xml_df

# -----------------------------
# Streamlit UI
# -----------------------------
st.subheader("1. Upload files")

print_logic_file = st.file_uploader("Print Logic report (.xlsx or .xls)", type=["xlsx", "xls"])
xml_zip_files = st.file_uploader("US Foods XML/email ZIP file(s)", type=["zip"], accept_multiple_files=True)
tracking_file = st.file_uploader("Tracking report to remove shipped jobs (.xlsx or .xls)", type=["xlsx", "xls"])
cancelled_file = st.file_uploader("Cancelled Status Report to highlight cancelled jobs (.xlsx or .xls)", type=["xlsx", "xls"])

st.subheader("2. Generate report")

if st.button("Generate US Foods Report", type="primary"):
    if not print_logic_file:
        st.error("Please upload the Print Logic report.")
    elif not xml_zip_files:
        st.error("Please upload at least one US Foods XML/email ZIP file.")
    else:
        try:
            report_bytes, final_df, xml_df = create_report(
                print_logic_file=print_logic_file,
                xml_zip_files=xml_zip_files,
                tracking_file=tracking_file,
                cancelled_file=cancelled_file,
            )

            output_name = f"US_Foods_Report_{datetime.now().strftime('%m%d%Y')}.xlsx"
            st.success(f"Report generated successfully. Final rows: {len(final_df)}")

            with st.expander("Preview final filtered rows"):
                st.dataframe(final_df, use_container_width=True)

            with st.expander("Preview extracted XML/email records"):
                st.dataframe(xml_df, use_container_width=True)

            st.download_button(
                label="Download Excel Report",
                data=report_bytes,
                file_name=output_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as e:
            st.error(f"Something went wrong: {e}")
            st.exception(e)

st.divider()
st.markdown("""
### Filter rule used by this app
The app keeps:
- Past ship dates
- Today's ship date
- The next upcoming future ship date only

The app removes:
- Any future ship date after the next upcoming ship date
""")
